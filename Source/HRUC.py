import vcrest, hrest, requests, json, logging, sys, os, pprint
from openpyxl import Workbook
from datetime import datetime
import re


# turns off security warnings for requests - probably not needed here since there are no direct requests.
requests.packages.urllib3.disable_warnings()

#configuring logging for this module - Comment this line when debugging is not needed
#logging.basicConfig(stream=sys.stderr, level=logging.DEBUG)
#logging.basicConfig(stream=sys.stderr, level=logging.ERROR)

############################################################################################
############################## GLOBAL VARS #################################################
############################################################################################
global HorizonHosts
global husername
global hpassword
global hdomain
workingDirectory= r""

############################################################################################
############################## FUNCTIONS  ##################################################
############################################################################################

def ParseArgs(argv):
    """
    Parses the data in the Config file passed in the command line for this module. Saves data to the Global Vars defined above.

    Parameters
    ----------
    argv : str
        The list of arguments passed to the parse args fxn. Argv should be a single-item list containing the path to the Config File
   
    Returns
    -------
    No Return - the Global Vars for the remainder of the program are populated.
        
    """
    contents = ''
    try:
        #should only be one argument
        if len(argv) != 1:
            print("Incorrect usage. Example: HRUC.py <ConfigFile>.ini")

        #determines if the argument is of the format *.ini
        match = re.search("\.ini$", argv[0])
        if match == None:
            print("Incorrect usage. Example: HRUC.py <ConfigFile>.ini")

        #should open the ini file and save its values to global vars
        with open(argv[0], "rt") as file_object:
            contents = file_object.read()
            #logging.debug(f"Printing the contents of the config file:{contents}")
    except:
        logging.debug("There was an issue parsing the args")
    
    if contents == '':
        logging.debug("There was an issue parsing the args - ending the program.")
        sys.exit()
    else:
        #Assuming contents has data, parse it as a JSON object
        params = json.loads(contents)
        logging.debug(f"Printing the JSON object {params}\n\n")

    if params != None:
        try:
            global HorizonHosts
            HorizonHosts = params['HorizonHosts']
            global husername
            husername = params['Username']
            global hpassword
            hpassword = params['Password']
            global hdomain
            hdomain = params['Domain']
        except KeyError:
            logging.error("The Config file is not configured correctly.")
            print("The config file has invalid/missing data. Ending the program.")
            sys.exit()
        except:
            logging.error("Uncaught exception parsing the config JSON.")
            print("There was an issue parsing the config file. Ending the program.")
            sys.exit()

    #printing the various data extracted from the config file
    logging.debug(f"HorizonHosts:{HorizonHosts},,username:{husername}\n\n")
    
def PopulateHorizonData(HorizonHosts,husername,hpassword,hdomain):
    """
    Takes in info for all Horizon Servers. Gets a sessionToken, PoolList, and vCenterInfo for each

    Parameters
    ----------
    HorizonHosts : List
        List of Horiozn Connection Server Hostnames
    husername : string
        String containing the domain username to login to Horizon
    hpassword : string
        String containing the plaintext password to login to Horizon
    hdomain : string 
        String containing the domain name to login to Horizon
    
    Returns
    -------
    HorizonServers : Dict
        A Dictionary of key: connectionServer, val:[AccessToken,Pools,vCenters]
        
    """
    HorizonServers = {} #should be - key:connectionServer, val:[AccessToken,Pools,vCenters]
    for i in range(len(HorizonHosts)):

        # Gets a session token for communicating with the Horizon Connection Server given by HorizonHosts[i]
        session = hrest.Connection.hv_connect(username=husername,password=hpassword,hostname=HorizonHosts[i],domain=hdomain)
        # Gets a list of dictionaries. Each dictionary corresponds to a pool in the Horizon Connection Server given by HorizonHosts[i]
        pools = hrest.Pools.list_hvpools(hostname=HorizonHosts[i], access_token=session)
        # Gets a list of dictionaries. Each dictionary corresponds to a vCenter Configuration in the Horizon Connection Server given by HorizonHosts[i]
        vCenters = hrest.Pools.get_VirtualCenters(hostname=HorizonHosts[i], access_token=session)
        vCList = []
        for vc in vCenters:
            vCList.append((vc['id'],vc['server_name']))
        HorizonServers[HorizonHosts[i]] = [session,pools,vCList]

    return HorizonServers

def PopulatevCenterData(HorizonServers, vusername, vpassword):
    """
    Takes in info the HorizonServers object. From this it will use the vCenter list to get vCenter data

    Parameters
    ----------
    HorizonServers : Dict
        A Dictionary of key: connectionServer, val:[AccessToken,Pools,vCenters] (vCenters is a list of tuples)
    vusername : string 
        A string containing the account used to log into vCenter
    vpassword : string
        A string containing the password used to log into vCenter
    
    Returns
    -------
    vCenterServers :  Dict[vchostname] = [finalClusters, vms]
        Dict with vCenterHostname as key, and value as a list [finalClsuters, vms]
            finalClusters is a list of lists - each inner list corresponds to a cluster in the vcenter hostname
                - inner list format: [friendlyname (vcenter-clustername), cluster id, cluster name, cpu capacity, cpuu usage,mem capacity,mem usage]
                - exa: ['10.173.8.2-Cluster-1', 'domain-c7', 'Cluster-1', 247860, 18406, 1765242, 316201]
            vms is a list of dictionaries corresponding to each VM in the vCenter
                - dictionary format: {'memory_size_MiB': <mem-sze>, 'vm': <vm-Id>, 'name': <vm display name>, 'power_state': <string power status>, 'cpu_count': <num vcpus>}
                - exa: {'memory_size_MiB': 4096, 'vm': 'vm-4226', 'name': 'cp-template-376a0896-fa6b-4a6f-b185-0449e1460222', 'power_state': 'POWERED_OFF', 'cpu_count': 2}
        
    """  
    vCenterServers = {} # dict of key:vcenterHostname, value: vmlist, clusters
 
    for key in HorizonServers: #iterates through each connection server
        vclist = HorizonServers[key][2] # this should be the the vcenter list for the connection server
        for y in range(len(vclist)):
            vcid,vchostname = vclist[y] # gets the id and hostname of the vcenter at index y
            # Get an auth session for vchostname
            session = vcrest.getvCenterSession(vchostname,vusername,vpassword)
            logging.debug(f"\n\n\n vCenter Server: id:{vcid}, hostname:{vchostname},session:{session}\n")
            
            #get all of the cluster data for this vcenter:
            clustersjson = json.loads((vcrest.GetClusters(session,vchostname)).text)
            clusters = clustersjson["value"]
            #logging.debug(f"               Clusters = {clusters}\n")

            #get all of the VM data for this vcenter 
            vmsjson = json.loads((vcrest.GetVMs(session,vchostname)).text)
            vms = vmsjson["value"] 
            #logging.debug(f"               VMs = {vms}\n")

            #returns a  a list of lists. Each sub-list corresponds to a cluster and  contains: [vCenterHostname, clusterID, CPU Capacity, CPU Used, Memory Capacity, Memory Used]
            clustersSDK = vcrest.getAllClusterInfoSingle(vchostname, vusername, vpassword)
            #logging.debug(f"               ClustersSDK = {clustersSDK}\n")
           
            if len(clustersSDK) != len (clusters):
                print("Error - the vCenter SDK and API are returning different sized cluster lists")
                sys.exit()
            
            #a list of lists = each inner list is [friendlyname (vcenter-clustername), cluster id, cluster name,cpuc,cpuu,memc,memu]
            finalClusters = []
            for i in range(len(clustersSDK)):
                for j in range(len(clusters)):
                    clsdkid = clustersSDK[i][1]
                    clid = clusters[j]['cluster']
                    if clid in clsdkid: #(exa. domain-c7 in 'vim.ClusterComputeResource:domain-c7')
                        logging.debug("Found a match for the two cluster IDs")
                        clname = clusters[j]['name']
                        cpuc = clustersSDK[i][2]
                        cpuu = clustersSDK[i][3]
                        memc = clustersSDK[i][4]
                        memu = clustersSDK[i][5]
                        friendlyName = vchostname + '-' + clname
                        templist = [friendlyName, clid, clname,cpuc,cpuu,memc,memu]
                        finalClusters.append(templist)
                        logging.debug(f"               FinalClusters = {finalClusters}\n")
            vCenterServers[vchostname] = [finalClusters, vms]

    return vCenterServers

def ParseData(HorizonServers, vCenterServers):
    """
    Takes in info for all Horizon Servers and vCenter Servers, and produces a PoolDict which contains a dictionary with an entry for every pool across all environments

    Parameters
    ----------
    HorizonServers : Dict
        A Dictionary of key: connectionServer, val:[AccessToken,Pools,vCenters]
    
     vCenterServers :  Dict[vchostname] = [finalClusters, vms]
        Dict with vCenterHostname as key, and value as a list [finalClsuters, vms]
            finalClusters is a list of lists - each inner list corresponds to a cluster in the vcenter hostname
                - inner list format: [friendlyname (vcenter-clustername), cluster id, cluster name, cpu capacity, cpuu usage,mem capacity,mem usage]
                - exa: ['10.173.8.2-Cluster-1', 'domain-c7', 'Cluster-1', 247860, 18406, 1765242, 316201]
            vms is a list of dictionaries corresponding to each VM in the vCenter
                - dictionary format: {'memory_size_MiB': <mem-sze>, 'vm': <vm-Id>, 'name': <vm display name>, 'power_state': <string power status>, 'cpu_count': <num vcpus>}
                - exa: {'memory_size_MiB': 4096, 'vm': 'vm-4226', 'name': 'cp-template-376a0896-fa6b-4a6f-b185-0449e1460222', 'power_state': 'POWERED_OFF', 'cpu_count': 2}

    Returns
    -------
    poolDict : dict 
        Dict with key='poolname' and value=dict{"PoolName", "vCenterID", "ClusterID", etc}
        
    """
    poolDict = {}
    #iterate through each Connection Server
        #iterate through each pool in each Connection Server
    for key in HorizonServers:
        poollist = HorizonServers[key][1]
            #list of tuples (vcid, vcenterIP/hostname)
        vclist = HorizonServers[key][2]
        logging.debug(f"Printing vclist for this horizon server = {vclist}")
        for z in range(len(poollist)):
            tempPool = poollist[z]
            #logging.debug(f"\n\nTempPool{key}{z} = {tempPool}\n\n")
            try:
                pattern_naming_settings = tempPool.get('pattern_naming_settings', {})
                provisioning_settings = tempPool.get('provisioning_settings',{})

                poolName = tempPool.get('name',"Unknown")
                logging.debug(f"Saving pool data to the poolDict for pool = {poolName}")
                poolDict[poolName] = {}
                poolDict[poolName]["PoolName"] = poolName
                poolDict[poolName]["vCenterID"] = tempPool.get('vcenter_id', "Unknown") #id of the vCenter this pool is in
                poolDict[poolName]["ClusterID"] = provisioning_settings.get('host_or_cluster_id', "Unknown") #id of the cluster
                poolDict[poolName]["ParentID"] = provisioning_settings.get('parent_vm_id', "Unknown")
                poolDict[poolName]["PoolStatus"] = tempPool.get('enabled', False) #pool status
                poolDict[poolName]["Provisioning"] = tempPool.get('enable_provisioning',False) #determines whether provisioning is currently on
                poolDict[poolName]["MaxMachines"] = pattern_naming_settings.get('max_number_of_machines', 0) # max number of machines
                poolDict[poolName]["MinMachines"] = pattern_naming_settings.get('min_number_of_machines', 0)
                poolDict[poolName]["SpareMachines"] = pattern_naming_settings.get('number_of_spare_machines', 0)
                poolDict[poolName]["NamingPattern"] = pattern_naming_settings.get('naming_pattern', "Unknown")
                poolDict[poolName]["Deleting"] = tempPool.get('delete_in_progress', False)
                poolDict[poolName]["Source"] = tempPool.get('image_source', "Unknown") # : 'VIRTUAL_CENTER'
                poolDict[poolName]["Type"] = tempPool.get('type', "Unknown")

            except Exception as e:
                logging.error(f"\n Likely KeyError: {tempPool['name']}") 
                logging.error(f"Exception = {e}\n")

            try: 
                #ITERATE THROUGH EACH vCENTER ASSOCIATED WITH THIS HORIZON SERVEr
                for h in range(len(vclist)):
                    #find the vcenter server name from this pools associated vcenter ID
                    if poolDict[poolName]["vCenterID"] == vclist[h][0]: #ID of the vcenter
                            #if tempvCList[h][1] is an IP, it will get mapped to DNS
                        poolDict[poolName]["vCenterServer"] = vclist[h][1]
            
                    # find vcenter clustername and friendly name for the current pool
        
                    clusterlist = vCenterServers[vclist[h][1]][0]
                    for p in range(len(clusterlist)):
                            # if the pool's clusterid == the clusterlist[p] id      
                        if (poolDict[poolName]["ClusterID"] == clusterlist[p][1]):
                            poolDict[poolName]["ClusterName"] = clusterlist[p][2]                      
                            poolDict[poolName]["VCCFriendlyName"] = clusterlist[p][0]
            except Exception as e: 
                logging.error("Unplanned error gettng the vCenter server from ID")
                logging.error(f"Exception = {e}")
  
            try: 
                vcName = poolDict[poolName].get("vCenterServer", "Unknown")
                if vcName == "Unknown":
                    poolDict[poolName]["VMName"] = "Unknown"
                    poolDict[poolName]["Memory"] = 0
                    poolDict[poolName]["CPU"] = 0
                    continue
                vmlist = vCenterServers[vcName][1]

                #logging.debug(f"Number of VMs in {vcName} is {len(vmlist)}")
                #iterating through the vmlist for this pools vcenter
                for p in range(len(vmlist)):
                   vmid = vmlist[p]['vm'] 
                   #logging.info(f"\n\n\n ID of VM whose name I'm trying to find = {vmid}")
                   if vmid == poolDict[poolName]["ParentID"]:
                        # if a vm list vm has the same id
                        poolDict[poolName]["VMName"] = vmlist[p]['name']
                        poolDict[poolName]["Memory"] = vmlist[p]['memory_size_MiB']
                        poolDict[poolName]["CPU"] = vmlist[p]['cpu_count']

                logging.debug(f"Printing the dict for {poolName} : {poolDict[poolName]} \n\n")
            except Exception as e:
                logging.error("\n\nSomething went wrong matching gold images to vcVMs")
                logging.error(f"Exception is {e}\n\n")
   

    pp = pprint.PrettyPrinter(indent=4)
    logging.info(f"\n\nPoolDict = {pp.pformat(poolDict)}")
    
    return poolDict

def CreateCodeDict(vCenterServers, poolDict):
    """
    Takes in info for all Horizon Pools and vCenter Servers, and produces a CodeDict which can be easily iterated through when writing the data to Excel

    Parameters
    ----------
    vCenterServers :  Dict[vchostname] = [finalClusters, vms]
        Dict with vCenterHostname as key, and value as a list [finalClsuters, vms]
            finalClusters is a list of lists - each inner list corresponds to a cluster in the vcenter hostname
                - inner list format: [friendlyname (vcenter-clustername), cluster id, cluster name, cpu capacity, cpuu usage,mem capacity,mem usage]
                - exa: ['10.173.8.2-Cluster-1', 'domain-c7', 'Cluster-1', 247860, 18406, 1765242, 316201]
            vms is a list of dictionaries corresponding to each VM in the vCenter
                - dictionary format: {'memory_size_MiB': <mem-sze>, 'vm': <vm-Id>, 'name': <vm display name>, 'power_state': <string power status>, 'cpu_count': <num vcpus>}
                - exa: {'memory_size_MiB': 4096, 'vm': 'vm-4226', 'name': 'cp-template-376a0896-fa6b-4a6f-b185-0449e1460222', 'power_state': 'POWERED_OFF', 'cpu_count': 2}
    PoolDict : Dict
            A Dictionary of key: PoolName, value=dict{"PoolName", "vCenterID", "ClusterID", etc}

    Returns
    -------
    codeDict : dict 
        Dict with key='VcenterHost-ClusterName' and value=list[PoolDictEntry, PoolDictEntry,...]
        This is easy to iterate through when writing to an excel file
        
    """
    codeDict = {}
    for val in vCenterServers.values():
        clusterlist = val[0]
        for r in range(len(clusterlist)):
            clustercode = clusterlist[r][0]
            templist = []
            for v in poolDict.values():
                if v.get("VCCFriendlyName","Unknown") == clustercode:
                    templist.append(v)
            codeDict[clustercode] = templist
    return codeDict



def WriteToExcel(vCenterServers, codeDict):
    """
    Takes in info for all Horizon Pools and vCenter Servers, and produces a CodeDict which can be easily iterated through when writing the data to Excel

    Parameters
    ----------
    vCenterServers :  Dict[vchostname] = [finalClusters, vms]
        Dict with vCenterHostname as key, and value as a list [finalClsuters, vms]
            finalClusters is a list of lists - each inner list corresponds to a cluster in the vcenter hostname
                - inner list format: [friendlyname (vcenter-clustername), cluster id, cluster name, cpu capacity, cpuu usage,mem capacity,mem usage]
                - exa: ['10.173.8.2-Cluster-1', 'domain-c7', 'Cluster-1', 247860, 18406, 1765242, 316201]
            vms is a list of dictionaries corresponding to each VM in the vCenter
                - dictionary format: {'memory_size_MiB': <mem-sze>, 'vm': <vm-Id>, 'name': <vm display name>, 'power_state': <string power status>, 'cpu_count': <num vcpus>}
                - exa: {'memory_size_MiB': 4096, 'vm': 'vm-4226', 'name': 'cp-template-376a0896-fa6b-4a6f-b185-0449e1460222', 'power_state': 'POWERED_OFF', 'cpu_count': 2}
    codeDict : dict 
        Dict with key='VcenterHost-ClusterName' and value=list[PoolDictEntry, PoolDictEntry,...]
        This is easy to iterate through when writing to an excel file

    Returns
    -------
    No return - an Excel file is outputted to the current working directory containing all of the previously captured data
        
    """

     # Initialize the Workbook
    wb = Workbook()

    # Create the first sheet: Data Sheet
    wb.create_sheet("Data Sheet",0)

    #delete the default sheet that gets created
    wb.remove(wb['Sheet'])
    

    index = 0
    for v in vCenterServers.values():
        clusterdata = v[0] 
        for g in range(len(clusterdata)):
            # Get the list of vCenter Clusters to use for the names of the sheets
            wb.create_sheet(clusterdata[g][0],index+1)
            index += 1

  
    for ws in wb.worksheets:
        logging.debug(f"###########Printing WS.Title = {ws.title}#####################")
        if ws.title == "Data Sheet":
            ws['A1'] = "Friendly Name"
            ws['B1'] = "vCenter Host"
            ws['C1'] = "CPU Capacity (GHz)"
            ws['D1'] = "CPU Usage (GHz)"
            ws['E1'] = "Memory Capacity (GB)"
            ws['F1'] = "Memory Usage (GB)"

            index = 0
            for v in vCenterServers.values():
                clusterdata = v[0] 
                for g in range(len(clusterdata)):
                    n = 2 + index
                    ws['A'+str(n)] = clusterdata[g][0] # Friednly NAME
                    ws['B'+str(n)] = clusterdata[g][0].split('-')[0] # vCENTER HOST
                    ws['C'+str(n)] = str(float(clusterdata[g][3])/1000) # = CPUc
                    ws['D'+str(n)] = str(float(clusterdata[g][4])/1000) # = CPUR
                    ws['E'+str(n)] = str(float(clusterdata[g][5])/1000) # = MemC
                    ws['F'+str(n)] = str(float(clusterdata[g][6])/1000) # = MemR  
                    index += 1
            continue # go to next sheet because we are done printing data

        logging.debug(f"Adding the headers for {ws.title}")
        ws['A1'] = "Friendly Name"
        ws['B1'] = "vCenterHost"
        ws['C1'] = "CPU Capacity (GHz)"
        ws['D1'] = "CPU Usage (GHz)"
        ws['E1'] = "Memory Capacity (GB)"
        ws['F1'] = "Memory Usage (GB)"
        

        for v in vCenterServers.values():
            clusterdata = v[0] 
            for i in range(len(clusterdata)):
                tempnum = i
                if ws.title == clusterdata[tempnum][0]:
                    logging.debug(f"Writing the data values for the header for {ws.title}")
                    ws['A2'] = clusterdata[tempnum][0]
                    ws['B2'] = clusterdata[tempnum][0].split('-')[0]
                    ws['C2'] = (float(clusterdata[tempnum][3])/1000)
                    ws['D2'] = (float(clusterdata[tempnum][4])/1000)
                    ws['E2'] = (float(clusterdata[tempnum][5])/1000)
                    ws['F2'] = (float(clusterdata[tempnum][6])/1000)

        
        # heading for the pool rows
        logging.debug(f"Writing the pool headers for {ws.title}")
        ws['A4'] = "Pool"
        ws['B4'] = "Max Machines"
        ws['C4'] = "Gold Image"
        ws['D4'] = "vCPUs"
        ws['E4'] = "vRAM (GB)"
        ws['F4'] = "Provisioing Status"
        ws['G4'] = "Pool Status"

        # Write the data for Each Row 
        temple = codeDict[ws.title] # returns a list of every pool as a dict with VCC = ws.title
        for i in range(len(temple)):
            try:
                if temple[i].get("Type", "UNDEFINED") != "AUTOMATED":
                    #it is an RDS Desktop pool or a Manual Pool, so skip it
                    continue
                if temple[i].get("PoolStatus", "UNDEFINED") == False:
                    ws['A'+str(i+5)] = temple[i].get("PoolName","UNDEFINED")
                    ws['B'+str(i+5)] = int(0)
                    ws['C'+str(i+5)] = temple[i].get("VMName", "DISABLED")
                    ws['D'+str(i+5)] = int(0)
                    ws['E'+str(i+5)] = int(0)
                    ws['F'+str(i+5)] = "DISABLED"
                    ws['G'+str(i+5)] = "DISABLED"
                else:
                    ws['A'+str(i+5)] = temple[i].get("PoolName","UNDEFINED")
                    ws['B'+str(i+5)] = int(temple[i].get("MaxMachines",0))
                    ws['C'+str(i+5)] = temple[i].get("VMName", "UNDEFINED")
                    ws['D'+str(i+5)] = float(temple[i].get("CPU",0))
                    ws['E'+str(i+5)] = (float(temple[i].get("Memory",0))/1000)
                    if temple[i].get("Provisioning", "UNDEFINED") == False:
                        ws['F'+str(i+5)] = "DISABLED"
                    else:
                        ws['F'+str(i+5)] = temple[i].get("Provisioning", "UNDEFINED")
                    ws['G'+str(i+5)] = temple[i].get("PoolStatus", "UNDEFINED")
                
            except KeyError as e:
                logging.debug(f"\n\nException = {e}")
                logging.debug(f"CodeDict KeyError: {temple[i]} \n\n")
               

        
      

        maxRow = 5 + (len(temple) - 1)
        ws['I4'] = "Totals"
        ws['I5'] = "Max Possible vCPUs (Enabled)"
        #ws['I6'] = "Max Possible GHz Usage (vCPUs*HostCoreFrequency) (Enabled)"
        ws['I7'] = "Max Possible vCPUs (All)"
        #ws['I8'] = "Max Possible GHz Usage (vCPUs*HostCoreFrequency) (All)"
        ws['I9'] = "Max Possible RAM Usage GB (Enabled) "
        ws['I10'] = "Max Possible RAM Usage GB (All)"
        ws['I11'] = "Max Possible RAM Usage % (Enabled) "
        ws['I12'] = "Max Possible RAM Usage % (All)"

                # Total vCPUS Enabled
        ws['J5'] = f"=SUMPRODUCT(D5:D{maxRow}*(F5:F{maxRow}=TRUE),B5:B{maxRow}*(F5:F{maxRow}=TRUE))"

                # Total vCPU multiplied by the GHz frequency of a single host's core (hardcoded the 2.3 GHz)  
        #ws['J6'] = f"=J5*2.3"

                # Total vCPUS
        ws['J7'] = f"=SUMPRODUCT(D5:D{maxRow},B5:B{maxRow})"

                # Total vCPU multiplied by the GHz frequency of a single host's core (hardcoded the 2.3 GHz)  
        #ws['J8'] = f"=J7*2.3"

                # Total Mem Enabled Sum of E columen (Mem) times B column (num machines) if the F Column is TRUE
        ws['J9'] = f"=SUMPRODUCT(E5:E{maxRow}*(F5:F{maxRow}=TRUE),B5:B{maxRow}*(F5:F{maxRow}=TRUE))"

                # Total Mem: Sum of E columen (Mem) times B column (num machines)
        ws['J10'] = f"=SUMPRODUCT(E5:E{maxRow},B5:B{maxRow})"

                # Total Mem Enabled Sum of E columen (Mem) times B column (num machines) if the F Column is TRUE
        ws['J11'] = f"=(SUMPRODUCT(E5:E{maxRow}*(F5:F{maxRow}=TRUE),B5:B{maxRow}*(F5:F{maxRow}=TRUE))/E2)*100"

                # Total Mem: Sum of E columen (Mem) times B column (num machines)
        ws['J12'] = f"=(SUMPRODUCT(E5:E{maxRow},B5:B{maxRow})/E2)*100"
       
    
    
    #### SAVE THE FILE WITH DATESTAMP IN THE NAME
    now = datetime.now()
    #formats the time string like 220915-174510
    dt_string = now.strftime("%Y%m%d-%H%M%S")
    wb.save(f"HorizonResourceUsage_{dt_string}.xlsx")
    logging.debug(f"printing current working directory: {os.getcwd()}")



#####################################################################################################################################################################
################################################################## MAIN #############################################################################################
#####################################################################################################################################################################

def main(argv):
    #if the working directory var is not empty, set the working directory
    if workingDirectory:
        os.chdir(workingDirectory)
    logging.debug(f"The current working directory of the file is : {str(os.getcwd())} ")

    #doesn't return aything. Fills the Global Vars from the Config File
    ParseArgs(argv)

    # Fills the HorizonServers variable with a dictionary containing: key:HorizonFQDN, val:[sessionToken, list[dicts(key=pool,val=pooldata)],list[tuple(vCenterHostname,vCenterID)]]
    global HorizonServers
    HorizonServers = PopulateHorizonData(HorizonHosts,husername,hpassword,hdomain)
    global vCenterServers
    vCenterServers = PopulatevCenterData(HorizonServers, husername, hpassword)

    # take in Horizon andd vCenter Data to create one unified data structure for use with the excelt fxn
    poolDict = ParseData(HorizonServers, vCenterServers)

    codeDict = CreateCodeDict(vCenterServers, poolDict)
    pp = pprint.PrettyPrinter(indent=4)
    logging.info(f"\n\n\n\nCodeDict = {pp.pformat(codeDict)}")

    WriteToExcel(vCenterServers, codeDict)

if __name__ == "__main__":
    #passes all the command line args to the main fxn
    #the usage of the script should be: uvh-inv.py <ConfigFile>.ini
    main(sys.argv[1:])


























